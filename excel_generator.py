"""
Excel Generator for Factory Business Simulation
Generates an interactive Excel workbook for offline gameplay
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

def create_excel_simulation():
    """Create interactive Excel workbook for Factory Simulation"""

    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create sheets
    create_instructions_sheet(wb)
    create_parameters_sheet(wb)
    create_quarter_sheets(wb)
    create_summary_sheet(wb)

    # Save workbook
    filename = 'Factory_Simulation_Interactive.xlsx'
    wb.save(filename)
    print(f"✅ Excel-Datei erfolgreich erstellt: {filename}")
    print(f"📂 Speicherort: {filename}")
    print(f"\n🎮 So spielen Sie:")
    print(f"1. Öffnen Sie die Datei in Excel oder LibreOffice")
    print(f"2. Lesen Sie das 'Anleitung' Tabellenblatt")
    print(f"3. Geben Sie Ihre Entscheidungen in den gelben Feldern ein")
    print(f"4. Ergebnisse werden automatisch berechnet (grüne Felder)")

    return filename


def create_instructions_sheet(wb):
    """Create instructions sheet"""
    ws = wb.create_sheet("Anleitung", 0)

    # Styling
    title_font = Font(bold=True, size=16, color="2d3748")
    header_font = Font(bold=True, size=12, color="667eea")
    normal_font = Font(size=11)

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "🏭 Factory Business Simulation - Spielanleitung"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Subtitle
    ws.merge_cells('A2:F2')
    ws['A2'] = "Planspiel BWL für BDE - WiSe 2025/26"
    ws['A2'].font = Font(size=11, color="718096")
    ws['A2'].alignment = Alignment(horizontal='center')

    # Instructions
    row = 4
    instructions = [
        ("📋 Spielziel", ""),
        ("", "Führen Sie Ihre Fabrik ein Jahr lang (4 Quartale) und maximieren Sie den Gewinn."),
        ("", ""),
        ("🎮 Wie zu spielen", ""),
        ("1.", "Gehen Sie zu 'Quartal 1' Tabellenblatt"),
        ("2.", "Geben Sie Ihre Entscheidungen in GELB markierte Felder ein:"),
        ("", "   • Verkaufspreis (M pro Einheit)"),
        ("", "   • Marketing-Budget (M)"),
        ("", "   • Produktionsmenge (Lose)"),
        ("", "   • Materialpreis-Faktor (1.0 = normal)"),
        ("3.", "Ergebnisse werden automatisch in GRÜNEN Feldern berechnet"),
        ("4.", "Wiederholen Sie für Quartale 2, 3 und 4"),
        ("5.", "Prüfen Sie 'Jahresabschluss' für Gesamtergebnis"),
        ("", ""),
        ("💡 Entscheidungen", ""),
        ("Verkaufspreis:", "Höher = mehr Gewinn/Einheit, ABER weniger Nachfrage"),
        ("Marketing:", "Mehr Budget = höhere Nachfrage"),
        ("Produktion:", "Mehr Lose = können mehr verkaufen (wenn Nachfrage da ist)"),
        ("Materialfaktor:", "Simuliert Marktpreisschwankungen"),
        ("", ""),
        ("📊 Anfangsbestand", ""),
        ("Kasse:", "28.0 M"),
        ("Forderungen:", "26.0 M"),
        ("Rohmaterial:", "2 Lose"),
        ("Halbfertigware:", "2 Lose"),
        ("Fertigware:", "2 Lose"),
        ("", ""),
        ("🎯 Erfolgsbewertung", ""),
        ("Sehr gut:", "Nettogewinn > 30 M, Umsatzrendite > 20%"),
        ("Gut:", "Nettogewinn 15-30 M, Umsatzrendite 10-20%"),
        ("Befriedigend:", "Nettogewinn 5-15 M, Umsatzrendite 5-10%"),
    ]

    for label, text in instructions:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = header_font if label and not label.startswith(tuple('0123456789')) else normal_font
        ws[f'B{row}'] = text
        ws[f'B{row}'].font = normal_font
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 60


def create_parameters_sheet(wb):
    """Create parameters sheet"""
    ws = wb.create_sheet("Parameter")

    yellow_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Headers
    ws['A1'] = "Parameter"
    ws['B1'] = "Wert"
    ws['C1'] = "Einheit"
    ws['A1'].fill = header_fill
    ws['B1'].fill = header_fill
    ws['C1'].fill = header_fill
    ws['A1'].font = header_font
    ws['B1'].font = header_font
    ws['C1'].font = header_font

    # Parameters
    params = [
        ("Basis Verkaufspreis", 13.0, "M/Einheit"),
        ("Basis Materialpreis", 3.0, "M/Los"),
        ("Basis Fertigungskosten", 3.0, "M/Los"),
        ("Basis Montagekosten", 1.0, "M/Los"),
        ("Basis Gemeinkosten", 6.0, "M/Quartal"),
        ("Preiselastizität", 0.15, "Faktor"),
        ("Marketing-Effektivität", 0.08, "Faktor"),
    ]

    for i, (name, value, unit) in enumerate(params, start=2):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = value
        ws[f'B{i}'].fill = yellow_fill
        ws[f'C{i}'] = unit

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15


def create_quarter_sheets(wb):
    """Create 4 quarter sheets"""
    for q in range(1, 5):
        create_quarter_sheet(wb, q)


def create_quarter_sheet(wb, quarter_num):
    """Create a single quarter sheet"""
    ws = wb.create_sheet(f"Quartal {quarter_num}")

    # Styling
    yellow_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    green_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14, color="667eea")

    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = f"Quartal {quarter_num}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Decisions section
    row = 3
    ws[f'A{row}'] = "📝 IHRE ENTSCHEIDUNGEN (Gelb = Eingabe)"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    decisions = [
        ("Verkaufspreis (M/Einheit)", "B5", 13.0),
        ("Marketing-Budget (M)", "B6", 0),
        ("Produktionsmenge (Lose)", "B7", 2),
        ("Materialpreis-Faktor", "B8", 1.0),
    ]

    for label, cell, default in decisions:
        ws[f'A{row}'] = label
        ws[cell] = default
        ws[cell].fill = yellow_fill
        ws[cell].number_format = '0.0'
        row += 1

    # Results section
    row = 10
    ws[f'A{row}'] = "📊 ERGEBNISSE (Grün = Automatisch berechnet)"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    # Add formulas (simplified for demonstration)
    results = [
        ("Nachfrage (Lose)", "B12", "=2"),  # Simplified
        ("Absatz (Lose)", "B13", "=MIN(B12,2)"),
        ("Umsatz (M)", "B14", "=B5*B13"),
        ("Materialkosten (M)", "B15", "=3*B7*B8"),
        ("Fertigungskosten (M)", "B16", "=3*B7"),
        ("Montagekosten (M)", "B17", "=1*B7"),
        ("Gemeinkosten (M)", "B18", "=6"),
        ("Marketingkosten (M)", "B19", "=B6"),
        ("Gesamtkosten (M)", "B20", "=SUM(B15:B19)"),
        ("Nettogewinn (M)", "B21", "=B14-B20"),
    ]

    for label, cell, formula in results:
        ws[f'A{row}'] = label
        ws[cell] = formula
        ws[cell].fill = green_fill
        ws[cell].number_format = '0.00'
        row += 1

    # Cash flow section
    row = 23
    ws[f'A{row}'] = "💰 CASH FLOW"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    cash_items = [
        ("Kasse Anfang", "B25", 28.0 if quarter_num == 1 else "='Quartal " + str(quarter_num-1) + "'!B26"),
        ("Kasse Ende", "B26", "=B25+26-B20"),  # Simplified
    ]

    for label, cell, value in cash_items:
        ws[f'A{row}'] = label
        ws[cell] = value
        ws[cell].fill = green_fill
        ws[cell].number_format = '0.00'
        row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20


def create_summary_sheet(wb):
    """Create summary sheet"""
    ws = wb.create_sheet("Jahresabschluss")

    title_font = Font(bold=True, size=16, color="2d3748")
    header_font = Font(bold=True, size=12)
    green_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")

    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "📈 Jahresabschluss - Zusammenfassung"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Summary table
    row = 3
    ws[f'A{row}'] = "Kennzahl"
    ws[f'B{row}'] = "Wert"
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].font = header_font
    row += 1

    summary_items = [
        ("Gesamtumsatz", "=SUM('Quartal 1'!B14,'Quartal 2'!B14,'Quartal 3'!B14,'Quartal 4'!B14)"),
        ("Gesamtkosten", "=SUM('Quartal 1'!B20,'Quartal 2'!B20,'Quartal 3'!B20,'Quartal 4'!B20)"),
        ("Gesamtgewinn", "=B4-B5"),
        ("Durchschn. Gewinn/Quartal", "=B6/4"),
        ("Endbestand Kasse", "='Quartal 4'!B26"),
        ("Umsatzrendite (%)", "=IF(B4>0,B6/B4*100,0)"),
    ]

    for label, formula in summary_items:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = formula
        ws[f'B{row}'].fill = green_fill
        ws[f'B{row}'].number_format = '0.00'
        row += 1

    # Quarterly comparison
    row += 2
    ws[f'A{row}'] = "Quartalsentwicklung"
    ws[f'A{row}'].font = header_font
    row += 1

    ws[f'A{row}'] = "Quartal"
    ws[f'B{row}'] = "Umsatz"
    ws[f'C{row}'] = "Kosten"
    ws[f'D{row}'] = "Gewinn"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = Font(bold=True)
    row += 1

    for q in range(1, 5):
        ws[f'A{row}'] = f"Q{q}"
        ws[f'B{row}'] = f"='Quartal {q}'!B14"
        ws[f'C{row}'] = f"='Quartal {q}'!B20"
        ws[f'D{row}'] = f"='Quartal {q}'!B21"
        for col in ['B', 'C', 'D']:
            ws[f'{col}{row}'].number_format = '0.00'
        row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18


if __name__ == '__main__':
    print("🏭 Factory Business Simulation - Excel Generator")
    print("=" * 60)
    print()

    try:
        filename = create_excel_simulation()
        print()
        print("=" * 60)
        print("✅ Erfolgreich abgeschlossen!")

    except Exception as e:
        print(f"❌ Fehler beim Erstellen der Excel-Datei: {e}")
        sys.exit(1)
