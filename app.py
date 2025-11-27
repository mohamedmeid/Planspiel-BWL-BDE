"""
Flask Web Application for Factory Business Simulation
Provides web-based interface for the game
Includes Abschreibungen, Zinsen, and Steuern
"""

from flask import Flask, render_template, request, jsonify, session, send_file
import json
import os
from datetime import datetime
from factory_simulator import FactorySimulator, GameParameters, QuarterResult
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'factory_simulation_secret_key_2025'

# Store simulators in memory (in production, use database)
simulators = {}


@app.route('/')
def index():
    """Main game interface"""
    return render_template('index.html')


@app.route('/api/start_game', methods=['POST'])
def start_game():
    """Initialize a new game"""
    data = request.json
    
    # Create game parameters
    params = GameParameters(
        base_sales_price=float(data.get('base_sales_price', 13.0)),
        base_material_price=float(data.get('base_material_price', 3.0)),
        base_production_cost=float(data.get('base_production_cost', 3.0)),
        base_assembly_cost=float(data.get('base_assembly_cost', 1.0)),
        base_overhead_cost=float(data.get('base_overhead_cost', 6.0))
    )
    
    # Create simulator
    game_id = data.get('game_id', 'default')
    simulator = FactorySimulator(params)
    simulators[game_id] = simulator
    
    return jsonify({
        'success': True,
        'game_id': game_id,
        'initial_state': {
            'cash': simulator.cash,
            'accounts_receivable': simulator.accounts_receivable,
            'raw_material_inventory': simulator.raw_material_inventory,
            'work_in_progress': simulator.work_in_progress,
            'finished_goods_inventory': simulator.finished_goods_inventory
        },
        'financial_params': {
            'depreciation_per_quarter': params.depreciation_per_quarter,
            'interest_per_quarter': params.interest_per_quarter,
            'tax_rate': params.tax_rate * 100
        }
    })


@app.route('/api/simulate_quarter', methods=['POST'])
def simulate_quarter():
    """Simulate one quarter with given decisions"""
    data = request.json
    game_id = data.get('game_id', 'default')
    
    if game_id not in simulators:
        return jsonify({'success': False, 'error': 'Game not found'}), 404
    
    simulator = simulators[game_id]
    
    # Extract decisions
    sales_price = float(data.get('sales_price', 13.0))
    marketing_budget = float(data.get('marketing_budget', 0))
    production_lots = int(data.get('production_lots', 2))
    material_purchase_lots = int(data.get('material_purchase_lots', 2))
    material_market_factor = float(data.get('material_market_factor', 1.0))
    overhead_factor = float(data.get('overhead_factor', 1.0))
    
    # Run simulation
    result = simulator.simulate_quarter(
        sales_price=sales_price,
        marketing_budget=marketing_budget,
        production_lots=production_lots,
        material_purchase_lots=material_purchase_lots,
        material_market_factor=material_market_factor,
        overhead_factor=overhead_factor
    )
    
    # Convert result to dict
    result_dict = {
        'quarter': result.quarter,
        'material_purchase_lots': result.material_purchase_lots,
        'production_lots': result.production_lots,
        'sales_price': result.sales_price,
        'sales_volume': result.sales_volume,
        'sales_revenue': result.sales_revenue,
        'material_cost': result.material_cost,
        'production_cost': result.production_cost,
        'assembly_cost': result.assembly_cost,
        'herstellungskosten': result.herstellungskosten,
        'overhead_cost': result.overhead_cost,
        'marketing_cost': result.marketing_cost,
        'depreciation': result.depreciation,
        'interest': result.interest,
        'total_operating_cost': result.total_operating_cost,
        'gross_profit': result.gross_profit,
        'ebit': result.ebit,
        'profit_before_tax': result.profit_before_tax,
        'tax': result.tax,
        'net_profit': result.net_profit,
        'raw_material_inventory': result.raw_material_inventory,
        'work_in_progress': result.work_in_progress,
        'finished_goods_inventory': result.finished_goods_inventory,
        'cash_beginning': result.cash_beginning,
        'cash_ending': result.cash_ending,
        'accounts_receivable': result.accounts_receivable
    }
    
    return jsonify({
        'success': True,
        'result': result_dict,
        'current_state': {
            'cash': simulator.cash,
            'accounts_receivable': simulator.accounts_receivable,
            'raw_material_inventory': simulator.raw_material_inventory,
            'work_in_progress': simulator.work_in_progress,
            'finished_goods_inventory': simulator.finished_goods_inventory
        }
    })


@app.route('/api/get_summary', methods=['GET'])
def get_summary():
    """Get game summary"""
    game_id = request.args.get('game_id', 'default')
    
    if game_id not in simulators:
        return jsonify({'success': False, 'error': 'Game not found'}), 404
    
    simulator = simulators[game_id]
    summary = simulator.get_summary()
    
    return jsonify({
        'success': True,
        'summary': summary,
        'results': [
            {
                'quarter': r.quarter,
                'sales_revenue': r.sales_revenue,
                'herstellungskosten': r.herstellungskosten,
                'gross_profit': r.gross_profit,
                'ebit': r.ebit,
                'profit_before_tax': r.profit_before_tax,
                'tax': r.tax,
                'net_profit': r.net_profit,
                'cash_ending': r.cash_ending
            }
            for r in simulator.results
        ]
    })


@app.route('/api/export_results', methods=['GET'])
def export_results():
    """Export game results as JSON"""
    game_id = request.args.get('game_id', 'default')

    if game_id not in simulators:
        return jsonify({'success': False, 'error': 'Game not found'}), 404

    simulator = simulators[game_id]
    filename = f"factory_results_{game_id}.json"
    simulator.export_results(filename)

    return jsonify({
        'success': True,
        'filename': filename,
        'message': 'Results exported successfully'
    })


@app.route('/api/export_excel', methods=['GET'])
def export_excel():
    """Export game results as a multi-sheet professional Excel report"""
    game_id = request.args.get('game_id', 'default')

    if game_id not in simulators:
        return jsonify({'success': False, 'error': 'Game not found'}), 404

    simulator = simulators[game_id]
    results = simulator.results
    summary = simulator.get_summary()
    params = simulator.params

    # Create workbook
    wb = Workbook()
    
    # Styles
    style_header = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    style_subheader = PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid")
    style_success = PatternFill(start_color="c6f6d5", end_color="c6f6d5", fill_type="solid")
    style_danger = PatternFill(start_color="fed7d7", end_color="fed7d7", fill_type="solid")
    
    font_title = Font(bold=True, size=16, color="2d3748")
    font_header = Font(bold=True, color="FFFFFF")
    font_bold = Font(bold=True)
    
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    def setup_header(ws, title, subtitle):
        ws.merge_cells('A1:F1')
        ws['A1'] = title
        ws['A1'].font = font_title
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:F2')
        ws['A2'] = subtitle
        ws['A2'].font = Font(italic=True, color="718096")
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A3:F3')
        ws['A3'] = f"TechGear Solutions GmbH - Report generiert am: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A3'].alignment = Alignment(horizontal='center')

    # ==========================================
    # SHEET 1: Management Summary
    # ==========================================
    ws_sum = wb.active
    ws_sum.title = "Management Summary"
    setup_header(ws_sum, "📊 Management Summary", "Wichtigste Kennzahlen auf einen Blick")
    
    # KPIs Table
    ws_sum['A5'] = "Finanz-Kennzahlen (Gesamtjahr)"
    ws_sum['A5'].font = Font(bold=True, size=12)
    
    kpis = [
        ("Gesamtumsatz", summary['total_revenue'], "M"),
        ("Reingewinn (Netto)", summary['total_net_profit'], "M"),
        ("Umsatzrendite (ROS)", summary['return_on_sales'], "%"),
        ("Endbestand Kasse", summary['final_cash'], "M"),
        ("Gesamte Steuern", summary['total_tax'], "M")
    ]
    
    row = 6
    for label, value, unit in kpis:
        ws_sum[f'A{row}'] = label
        ws_sum[f'B{row}'] = value
        ws_sum[f'B{row}'].number_format = f'0.00 "{unit}"'
        ws_sum[f'B{row}'].font = font_bold
        
        # Color coding for Profit and Cash
        if "Netto" in label or "Kasse" in label:
             ws_sum[f'B{row}'].fill = style_success if value >= 0 else style_danger
             
        row += 1

    ws_sum.column_dimensions['A'].width = 25
    ws_sum.column_dimensions['B'].width = 15

    # ==========================================
    # SHEET 2: GuV Detail
    # ==========================================
    ws_guv = wb.create_sheet("GuV Detail")
    setup_header(ws_guv, "📉 Gewinn- und Verlustrechnung", "Detaillierte Aufstellung nach Quartalen")
    
    headers = ['Position', 'Q1', 'Q2', 'Q3', 'Q4', 'GESAMT']
    for col, h in enumerate(headers, 1):
        cell = ws_guv.cell(row=5, column=col, value=h)
        cell.fill = style_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal='center')
    
    # Data Rows
    guv_rows = [
        ("Umsatzerlöse", "sales_revenue", False),
        ("Herstellungskosten", "herstellungskosten", True), # True = negative
        ("= Bruttoergebnis", "gross_profit", False, True), # True = bold row
        ("Gemeinkosten", "overhead_cost", True),
        ("Marketing", "marketing_cost", True),
        ("Abschreibungen", "depreciation", True),
        ("= EBIT", "ebit", False, True),
        ("Zinsen", "interest", True),
        ("= Gewinn vor Steuern", "profit_before_tax", False, True),
        ("Steuern", "tax", True),
        ("= Gewinn nach Steuern", "net_profit", False, True)
    ]
    
    current_row = 6
    for label, attr, is_expense, *is_bold in guv_rows:
        ws_guv.cell(row=current_row, column=1, value=label).font = font_bold if is_bold else None
        
        total_val = 0
        for i, r in enumerate(results):
            val = getattr(r, attr)
            total_val += val
            
            # Flip sign for expenses for visual representation? 
            # Standard GuV usually lists positive numbers but subtracts them.
            # Let's keep them positive but maybe red text?
            # Or use negative numbers as in the app.py logic. 
            # The app.py logic used negative numbers for visual clarity.
            
            display_val = -val if is_expense else val
            
            c = ws_guv.cell(row=current_row, column=i+2, value=display_val)
            c.number_format = '0.00 "M"'
            if is_bold: c.font = font_bold
            if is_expense: c.font = Font(color="C00000")
            if "Brutto" in label or "EBIT" in label or "Gewinn" in label:
                c.fill = style_subheader
        
        # Total Column
        total_display = -total_val if is_expense else total_val
        c_total = ws_guv.cell(row=current_row, column=6, value=total_display)
        c_total.number_format = '0.00 "M"'
        c_total.font = font_bold
        c_total.border = Border(left=Side(style='double'))
        
        current_row += 1

    ws_guv.column_dimensions['A'].width = 30
    for i in range(2, 7): ws_guv.column_dimensions[get_column_letter(i)].width = 15

    # ==========================================
    # SHEET 3: Cashflow & Bilanz
    # ==========================================
    ws_bal = wb.create_sheet("Cashflow & Bilanz")
    setup_header(ws_bal, "💰 Cashflow & Vermögenswerte", "Liquiditätsrechnung und Bestandsbewertung")
    
    # Cashflow Headers
    ws_bal['A5'] = "CASHFLOW RECHNUNG"
    ws_bal['A5'].font = Font(bold=True, size=12, color="667eea")
    
    headers = ['Position', 'Q1', 'Q2', 'Q3', 'Q4']
    for col, h in enumerate(headers, 1):
        ws_bal.cell(row=6, column=col, value=h).font = font_bold
        ws_bal.cell(row=6, column=col).border = Border(bottom=Side(style='medium'))

    # Cashflow Data Construction
    # Recalculate explicit flows for clarity
    cf_rows = [
        "Anfangsbestand Kasse",
        "+ Einzahlungen (Forderungen)",
        "- Ausz. Operativ (Mat/Prod/Gemein/Mark)",
        "- Ausz. Finanzen (Zinsen/Steuern)",
        "= Endbestand Kasse"
    ]
    
    r_idx = 7
    for label in cf_rows:
        ws_bal.cell(row=r_idx, column=1, value=label)
        r_idx += 1
        
    for i, res in enumerate(results):
        col = i + 2
        # Start
        ws_bal.cell(row=7, column=col, value=res.cash_beginning).number_format = '0.00'
        
        # In (Receivables from prev quarter)
        inflow = res.cash_beginning + res.sales_revenue - res.cash_ending + res.total_operating_cost # Reverse eng? 
        # Actually simpler: Inflow = Receivables collected. 
        # logic: cash_end = cash_start + receivables_collected - total_cash_costs
        # receivables_collected = cash_end - cash_start + total_cash_costs
        # Wait, simpler: The simulator logic says "Cash in: customer payments (previous quarter receivables)".
        # In Q1, we collect initial receivables.
        
        # Let's recalculate based on simulator logic:
        # cash_in = self.accounts_receivable (from PREVIOUS state).
        # We don't have previous state easily in 'results' list alone without looking back.
        # But we can infer: inflow = (cash_ending - cash_beginning) + total_cash_costs
        
        cash_costs_op = (res.material_cost + res.production_cost + res.assembly_cost + 
                        res.overhead_cost + res.marketing_cost)
        cash_costs_fin = res.interest + res.tax
        total_out = cash_costs_op + cash_costs_fin
        
        inflow = (res.cash_ending - res.cash_beginning) + total_out
        
        ws_bal.cell(row=8, column=col, value=inflow).number_format = '0.00'
        ws_bal.cell(row=9, column=col, value=-cash_costs_op).number_format = '0.00'
        ws_bal.cell(row=10, column=col, value=-cash_costs_fin).number_format = '0.00'
        
        c_end = ws_bal.cell(row=11, column=col, value=res.cash_ending)
        c_end.number_format = '0.00'
        c_end.font = font_bold
        c_end.fill = style_subheader

    # Asset Valuation (Inventory)
    r_start = 14
    ws_bal[f'A{r_start}'] = "VERMÖGENSWERTE (Indikativ)"
    ws_bal[f'A{r_start}'].font = Font(bold=True, size=12, color="667eea")
    
    # Valuation Logic
    # Raw = Base Price (3.0)
    # WIP = Mat + Prod (3.0 + 3.0 = 6.0)
    # Finished = Mat + Prod + Assembly (3.0 + 3.0 + 1.0 = 7.0)
    val_raw = params.base_material_price
    val_wip = params.base_material_price + params.base_production_cost
    val_fin = params.base_material_price + params.base_production_cost + params.base_assembly_cost
    
    asset_rows = ["Liquide Mittel", "Forderungen (aus Verkauf)", "Vorräte (Bewertet)", "SUMME UMLAUFVERMÖGEN"]
    for i, l in enumerate(asset_rows):
        ws_bal.cell(row=r_start+1+i, column=1, value=l)

    for i, res in enumerate(results):
        col = i + 2
        # Cash
        ws_bal.cell(row=r_start+1, column=col, value=res.cash_ending).number_format = '0.00'
        # Receivables
        ws_bal.cell(row=r_start+2, column=col, value=res.accounts_receivable).number_format = '0.00'
        
        # Inventory Value
        inv_val = (res.raw_material_inventory * val_raw) + \
                  (res.work_in_progress * val_wip) + \
                  (res.finished_goods_inventory * val_fin)
        ws_bal.cell(row=r_start+3, column=col, value=inv_val).number_format = '0.00'
        
        # Sum
        total_assets = res.cash_ending + res.accounts_receivable + inv_val
        c_sum = ws_bal.cell(row=r_start+4, column=col, value=total_assets)
        c_sum.number_format = '0.00'
        c_sum.font = font_bold
        c_sum.border = Border(top=Side(style='thin'), bottom=Side(style='double'))

    ws_bal.column_dimensions['A'].width = 35
    
    # ==========================================
    # SHEET 4: Produktion & Lager
    # ==========================================
    ws_prod = wb.create_sheet("Produktion & Lager")
    setup_header(ws_prod, "🏭 Produktion & Logistik", "Mengenströme und Lagerbestände")
    
    headers = ['Kennzahl', 'Q1', 'Q2', 'Q3', 'Q4']
    for col, h in enumerate(headers, 1):
        ws_prod.cell(row=5, column=col, value=h).font = font_bold
        ws_prod.cell(row=5, column=col).fill = style_subheader

    prod_data = [
        ("MENGENSTRÖME", "", ""),
        ("Einkauf (Lose)", "material_purchase_lots", "0"),
        ("Produktion (Lose)", "production_lots", "0"),
        ("Absatz (Lose)", "sales_volume", "0"),
        ("", "", ""),
        ("LAGERBESTÄNDE (Ende)", "", ""),
        ("Rohmaterial", "raw_material_inventory", "0"),
        ("Halbfertigware (WIP)", "work_in_progress", "0"),
        ("Fertigware", "finished_goods_inventory", "0"),
        ("", "", ""),
        ("MARKT-DATEN", "", ""),
        ("Verkaufspreis", "sales_price", "0.00"),
        ("Marketing-Budget", "marketing_cost", "0.00")
    ]
    
    curr_row = 6
    for label, attr, fmt in prod_data:
        cell = ws_prod.cell(row=curr_row, column=1, value=label)
        if attr == "": # Section Header
            cell.font = Font(bold=True, color="667eea")
        else:
            for i, res in enumerate(results):
                val = getattr(res, attr)
                c = ws_prod.cell(row=curr_row, column=i+2, value=val)
                c.number_format = fmt
                c.alignment = Alignment(horizontal='center')
        curr_row += 1

    ws_prod.column_dimensions['A'].width = 30

    # Save file
    exports_dir = os.path.join(os.getcwd(), 'exports')
    os.makedirs(exports_dir, exist_ok=True)

    filename = f"TechGear_Report_{game_id}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    filepath = os.path.join(exports_dir, filename)
    wb.save(filepath)

    return send_file(filepath, as_attachment=True, download_name=filename)


if __name__ == '__main__':
    # Create templates directory if it doesn't exist
    os.makedirs('templates', exist_ok=True)
    os.makedirs('static', exist_ok=True)
    os.makedirs('exports', exist_ok=True)

    # Use environment variable for port (for cloud deployment)
    port = int(os.environ.get('PORT', 5001))
    debug = os.environ.get('FLASK_ENV') != 'production'

    app.run(debug=debug, host='0.0.0.0', port=port)
